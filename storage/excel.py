import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from models import Job

_OUTPUT_DIR = "output"

_HEADERS = [
    "#", "Title", "Company", "Location", "Remote",
    "Salary", "Tech Stack", "Job Type", "Source", "Posted At", "Apply URL",
]

_COL_WIDTHS = [4, 42, 26, 28, 8, 24, 36, 12, 16, 14, 60]


def export_to_excel(jobs: list[Job]) -> str:
    os.makedirs(_OUTPUT_DIR, exist_ok=True)
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    path = os.path.join(_OUTPUT_DIR, f"jobs_{timestamp}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="4472C4")
    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(wrap_text=True, vertical="top")

    for col_idx, (header, width) in enumerate(zip(_HEADERS, _COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 20

    alt_fill = PatternFill(fill_type="solid", fgColor="DCE6F1")

    for row_idx, job in enumerate(jobs, 2):
        fill = alt_fill if row_idx % 2 == 0 else None
        salary = job.salary_raw or (
            f"${job.salary_min:,}+" if job.salary_min else ""
        )
        tech = ", ".join(job.tech_stack[:8]) if job.tech_stack else ""
        values = [
            row_idx - 1,
            job.title,
            job.company,
            job.location,
            "Yes" if job.is_remote else "No",
            salary,
            tech,
            job.job_type or "",
            job.source,
            job.posted_at or "",
            job.apply_url,
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = wrap
            if fill:
                cell.fill = fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_HEADERS))}1"

    wb.save(path)
    return path
